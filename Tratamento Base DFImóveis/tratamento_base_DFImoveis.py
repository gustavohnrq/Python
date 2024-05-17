# -*- coding: utf-8 -*-

def funcao_aguas_claras():
    
    ################### MEDIAS DO GRUPO COM VAGA ##################
    
    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Apartamento com bairro "AGUAS CLARAS COM VAGA"
    filtro_aguas_claras = ((df["bairro"] == "AREAL") | (df["bairro"] == "NORTE")| (df["bairro"] == "SUL")) & (df["tipo"] == "Apartamento") & (df["oferta"] == "Venda") & (df["vagas"].notnull())
    filtro_aguas_claras = df[filtro_aguas_claras]

    # Remover valores absurdos
    media_valor_m2 = filtro_aguas_claras["valor_m2"].mean()
    limite_superior = media_valor_m2 * 2.0
    limite_inferior = media_valor_m2 * 0.5
    filtro_aguas_claras = filtro_aguas_claras[(filtro_aguas_claras["valor_m2"] >= limite_inferior) & (filtro_aguas_claras["valor_m2"] <= limite_superior)]

    # Clusterização
    X = filtro_aguas_claras[["valor_m2"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    filtro_aguas_claras["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = filtro_aguas_claras.groupby("cluster")["valor_m2"].mean().sort_values().index
    filtro_aguas_claras["cluster"] = filtro_aguas_claras["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    filtro_aguas_claras["quartos_group"] = pd.cut(filtro_aguas_claras["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 50, 75, 90, 130, 160, 200, np.inf]
    labels = ["<50", "50-75", "75-90", "90-130", "130-160", "160-200", ">200"]
    filtro_aguas_claras["grupo_metragem"] = pd.cut(filtro_aguas_claras["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "Areal": ["AREAL"],
        "Águas Claras Norte": ["NORTE"],
        "Águas Claras Sul": ["SUL"]
    }

    # Calculate averages by cluster
    media_clusters = filtro_aguas_claras.groupby("cluster")["valor_m2"].mean()

    # Calculate averages by group of area
    media_metragem = filtro_aguas_claras.groupby("grupo_metragem")["valor_m2"].mean()

    # Calculate averages by number of rooms
    media_quartos = filtro_aguas_claras.groupby("quartos_group")["valor_m2"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = filtro_aguas_claras[filtro_aguas_claras["bairro"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["valor_m2"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/AGUAS CLARAS COM VAGA.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)


    ################### MEDIAS DO GRUPO SEM VAGA ##################


    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Apartamento com bairro "AGUAS CLARAS SEM VAGA"
    filtro_aguas_claras = ((df["bairro"] == "AREAL") | (df["bairro"] == "NORTE")| (df["bairro"] == "SUL")) & (df["tipo"] == "Apartamento") & (df["oferta"] == "Venda") & (df["vagas"].isnull())
    filtro_aguas_claras = df[filtro_aguas_claras]

    # Remover valores absurdos
    media_valor_m2 = filtro_aguas_claras["valor_m2"].mean()
    limite_superior = media_valor_m2 * 2.0
    limite_inferior = media_valor_m2 * 0.5
    filtro_aguas_claras = filtro_aguas_claras[(filtro_aguas_claras["valor_m2"] >= limite_inferior) & (filtro_aguas_claras["valor_m2"] <= limite_superior)]

    # Clusterização
    X = filtro_aguas_claras[["valor_m2"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    filtro_aguas_claras["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = filtro_aguas_claras.groupby("cluster")["valor_m2"].mean().sort_values().index
    filtro_aguas_claras["cluster"] = filtro_aguas_claras["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    filtro_aguas_claras["quartos_group"] = pd.cut(filtro_aguas_claras["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 50, 75, 90, 130, 160, 200, np.inf]
    labels = ["<50", "50-75", "75-90", "90-130", "130-160", "160-200", ">200"]
    filtro_aguas_claras["grupo_metragem"] = pd.cut(filtro_aguas_claras["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "Areal": ["AREAL"],
        "Águas Claras Norte": ["NORTE"],
        "Águas Claras Sul": ["SUL"]
    }

    # Calculate averages by cluster
    media_clusters = filtro_aguas_claras.groupby("cluster")["valor_m2"].mean()

    # Calculate averages by group of area
    media_metragem = filtro_aguas_claras.groupby("grupo_metragem")["valor_m2"].mean()

    # Calculate averages by number of rooms
    media_quartos = filtro_aguas_claras.groupby("quartos_group")["valor_m2"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = filtro_aguas_claras[filtro_aguas_claras["bairro"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["valor_m2"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/AGUAS CLARAS SEM VAGA.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)

    ################### ORGANIZA OS RESULTADOS ##################


    # Organiza os rsultados

    import openpyxl

    # Carregar os arquivos Excel originais
    input_file1 = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/AGUAS CLARAS COM VAGA.xlsx"
    input_file2 = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/AGUAS CLARAS SEM VAGA.xlsx"

    workbook1 = openpyxl.load_workbook(input_file1)
    workbook2 = openpyxl.load_workbook(input_file2)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/AGUAS CLARAS.xlsx"
    result_workbook = openpyxl.Workbook()

    # Para cada aba no primeiro arquivo, copiar os resultados para o novo arquivo
    for sheet_name in workbook1.sheetnames:
        original_sheet = workbook1[sheet_name]
        result_sheet = result_workbook.create_sheet(title=sheet_name)

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

    # Para cada aba no segundo arquivo, copiar os resultados para o novo arquivo
    for sheet_name in workbook2.sheetnames:
        original_sheet = workbook2[sheet_name]
        result_sheet = result_workbook.create_sheet(title=sheet_name)

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

    # Salvar o novo arquivo Excel com os resultados finais
    result_workbook.save(output_file)
    print("Planilhas unidas e salvas em:", output_file)


    import os
    import openpyxl

    # Caminho do arquivo Excel com os resultados
    input_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/AGUAS CLARAS.xlsx"

    # Carregar o arquivo Excel original
    workbook = openpyxl.load_workbook(input_file, data_only=True)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Organizados/AGUAS CLARAS.xlsx"
    result_workbook = openpyxl.Workbook()

    # Remover a aba padrão criada automaticamente
    default_sheet = result_workbook.active
    result_workbook.remove(default_sheet)

    # Criar uma aba para os resultados finais
    result_sheet = result_workbook.create_sheet(title="AGUAS CLARAS")

    # Para cada aba no arquivo original, copiar os resultados para a nova aba no arquivo final
    for sheet_name in workbook.sheetnames:
        original_sheet = workbook[sheet_name]

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

        result_sheet.append([])  # Adicionar duas linhas de espaço entre as tabelas

    # Salvar o novo arquivo Excel com os resultados finais no mesmo local do arquivo de origem
    result_workbook.save(output_file)
    print("Resultados finais salvos em:", output_file)

    # Organiza os rsultados

    import os
    import openpyxl

    # Caminho do arquivo Excel com os resultados
    input_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/AGUAS CLARAS.xlsx"

    # Carregar o arquivo Excel original
    workbook = openpyxl.load_workbook(input_file, data_only=True)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Organizados/AGUAS CLARAS.xlsx"
    result_workbook = openpyxl.Workbook()

    # Remover a aba padrão criada automaticamente
    default_sheet = result_workbook.active
    result_workbook.remove(default_sheet)

    # Criar uma aba para os resultados finais
    result_sheet = result_workbook.create_sheet(title="AGUAS CLARAS")

    # Para cada aba no arquivo original, copiar os resultados para a nova aba no arquivo final
    for sheet_name in workbook.sheetnames:
        original_sheet = workbook[sheet_name]

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

        result_sheet.append([])  # Adicionar duas linhas de espaço entre as tabelas

    # Salvar o novo arquivo Excel com os resultados finais no mesmo local do arquivo de origem
    result_workbook.save(output_file)
    print("Resultados finais salvos em:", output_file)

def funcao_arniqueira():
    
        ################### MEDIAS ##################
    
    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Casa com bairro "ARNIQUEIRAS"
    filtro_arniqueira = (df["bairro"] == "ARNIQUEIRAS") & (df["tipo"] == "Casa") & (df["oferta"] == "Venda")
    df_arniqueira = df[filtro_arniqueira]

    # Remover valores absurdos
    media_valor_m2 = df_arniqueira["valor_m2"].mean()
    limite_superior = media_valor_m2 * 2.0
    limite_inferior = media_valor_m2 * 0.5
    df_arniqueira = df_arniqueira[(df_arniqueira["valor_m2"] >= limite_inferior) & (df_arniqueira["valor_m2"] <= limite_superior)]

    # Clusterização
    X = df_arniqueira[["valor_m2"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    df_arniqueira["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = df_arniqueira.groupby("cluster")["valor_m2"].mean().sort_values().index
    df_arniqueira["cluster"] = df_arniqueira["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    df_arniqueira["quartos_group"] = pd.cut(df_arniqueira["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 100, 200, 300, 400, 600, 800, np.inf]
    labels = ["<100", "100-200", "200-300", "300-400", "400-600", "600-800", ">800"]
    df_arniqueira["grupo_metragem"] = pd.cut(df_arniqueira["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "ARNIQUEIRA": [" "],
    }

    # Calculate averages by cluster
    media_clusters = df_arniqueira.groupby("cluster")["valor_m2"].mean()

    # Calculate averages by group of area
    media_metragem = df_arniqueira.groupby("grupo_metragem")["valor_m2"].mean()

    # Calculate averages by number of rooms
    media_quartos = df_arniqueira.groupby("quartos_group")["valor_m2"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = df_arniqueira[df_arniqueira["quadra"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["valor_m2"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/ARNIQUEIRAS.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Resultados_Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)

    ################### ORGANIZA OS RESULTADOS ##################

    import os
    import openpyxl

    # Caminho do arquivo Excel com os resultados
    input_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/ARNIQUEIRAS.xlsx"

    # Carregar o arquivo Excel original
    workbook = openpyxl.load_workbook(input_file, data_only=True)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Organizados/ARNIQUEIRAS.xlsx"
    result_workbook = openpyxl.Workbook()

    # Remover a aba padrão criada automaticamente
    default_sheet = result_workbook.active
    result_workbook.remove(default_sheet)

    # Criar uma aba para os resultados finais
    result_sheet = result_workbook.create_sheet(title="ARNIQUEIRAS")

    # Para cada aba no arquivo original, copiar os resultados para a nova aba no arquivo final
    for sheet_name in workbook.sheetnames:
        original_sheet = workbook[sheet_name]

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

        result_sheet.append([])  # Adicionar duas linhas de espaço entre as tabelas

    # Salvar o novo arquivo Excel com os resultados finais no mesmo local do arquivo de origem
    result_workbook.save(output_file)
    print("Resultados finais salvos em:", output_file)

def funcao_asa_norte():
        ################### MEDIAS DO GRUPO COM VAGA ##################
    
    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Apartamento com bairro "ASA NORTE COM VAGA"
    filtro_asa_norte = (df["bairro"] == "ASA NORTE") & (df["tipo"] == "Apartamento") & (df["oferta"] == "Venda") & (df["vagas"].notnull())
    filtro_asa_norte = df[filtro_asa_norte]

    # Remover valores absurdos
    media_valor_m2 = filtro_asa_norte["valor_m2"].mean()
    limite_superior = media_valor_m2 * 2.0
    limite_inferior = media_valor_m2 * 0.5
    filtro_asa_norte = filtro_asa_norte[(filtro_asa_norte["valor_m2"] >= limite_inferior) & (filtro_asa_norte["valor_m2"] <= limite_superior)]

    # Clusterização
    X = filtro_asa_norte[["valor_m2"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    filtro_asa_norte["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = filtro_asa_norte.groupby("cluster")["valor_m2"].mean().sort_values().index
    filtro_asa_norte["cluster"] = filtro_asa_norte["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    filtro_asa_norte["quartos_group"] = pd.cut(filtro_asa_norte["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 50, 75, 90, 130, 160, 200, np.inf]
    labels = ["<50", "50-75", "75-90", "90-130", "130-160", "160-200", ">200"]
    filtro_asa_norte["grupo_metragem"] = pd.cut(filtro_asa_norte["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "SQN 100": ["SQN 102", "SQN 103", "SQN 104", "SQN 105", "SQN 106", "SQN 107", "SQN 108", "SQN 109", "SQN 110", "SQN 111", "SQN 112", "SQN 113", "SQN 114", "SQN 115", "SQN 116"],
        "SQN 200": ["SQN 202", "SQN 203", "SQN 204", "SQN 205", "SQN 206", "SQN 207", "SQN 208", "SQN 209", "SQN 210", "SQN 211", "SQN 212", "SQN 213", "SQN 214", "SQN 215", "SQN 216"],
        "SQN 300": ["SQN 302", "SQN 303", "SQN 304", "SQN 305", "SQN 306", "SQN 307", "SQN 308", "SQN 309", "SQN 310", "SQN 311", "SQN 312", "SQN 313", "SQN 314", "SQN 315", "SQN 316"],
        "SQN 400": ["SQN 402", "SQN 403", "SQN 404", "SQN 405", "SQN 406", "SQN 407", "SQN 408", "SQN 409", "SQN 410", "SQN 411", "SQN 412", "SQN 413", "SQN 414", "SQN 415", "SQN 416"]
    }

    quadras_individuais = [
        "SQN 102", "SQN 103", "SQN 104", "SQN 105", "SQN 106", "SQN 107",
        "SQN 108", "SQN 109", "SQN 110", "SQN 111", "SQN 112", "SQN 113",
        "SQN 114", "SQN 115", "SQN 116", "SQN 202", "SQN 203", "SQN 204",
        "SQN 205", "SQN 206", "SQN 207", "SQN 208", "SQN 209", "SQN 210",
        "SQN 211", "SQN 212", "SQN 213", "SQN 214", "SQN 215", "SQN 216",
        "SQN 302", "SQN 303", "SQN 304", "SQN 305", "SQN 306", "SQN 307",
        "SQN 308", "SQN 309", "SQN 310", "SQN 311", "SQN 312", "SQN 313",
        "SQN 314", "SQN 315", "SQN 316", "SQN 402", "SQN 403", "SQN 404",
        "SQN 405", "SQN 406", "SQN 407", "SQN 408", "SQN 409", "SQN 410",
        "SQN 411", "SQN 412", "SQN 413", "SQN 414", "SQN 415", "SQN 416"
    ]

    for quadra in quadras_individuais:
        groups[quadra] = [quadra]

    # Calculate averages by cluster
    media_clusters = filtro_asa_norte.groupby("cluster")["valor_m2"].mean()

    # Calculate averages by group of area
    media_metragem = filtro_asa_norte.groupby("grupo_metragem")["valor_m2"].mean()

    # Calculate averages by number of rooms
    media_quartos = filtro_asa_norte.groupby("quartos_group")["valor_m2"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = filtro_asa_norte[filtro_asa_norte["quadra"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["valor_m2"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/ASA NORTE COM VAGA.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)

    ################### MEDIAS DO GRUPO SEM VAGA ##################

    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Apartamento com bairro "ASA NORTE SEM VAGA"
    filtro_asa_norte = (df["bairro"] == "ASA NORTE") & (df["tipo"] == "Apartamento") & (df["oferta"] == "Venda") & (df["vagas"].isnull())
    filtro_asa_norte = df[filtro_asa_norte]

    # Remover valores absurdos
    media_valor_m2 = filtro_asa_norte["valor_m2"].mean()
    limite_superior = media_valor_m2 * 2.0
    limite_inferior = media_valor_m2 * 0.5
    filtro_asa_norte = filtro_asa_norte[(filtro_asa_norte["valor_m2"] >= limite_inferior) & (filtro_asa_norte["valor_m2"] <= limite_superior)]

    # Clusterização
    X = filtro_asa_norte[["valor_m2"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    filtro_asa_norte["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = filtro_asa_norte.groupby("cluster")["valor_m2"].mean().sort_values().index
    filtro_asa_norte["cluster"] = filtro_asa_norte["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    filtro_asa_norte["quartos_group"] = pd.cut(filtro_asa_norte["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 50, 75, 90, 130, 160, 200, np.inf]
    labels = ["<50", "50-75", "75-90", "90-130", "130-160", "160-200", ">200"]
    filtro_asa_norte["grupo_metragem"] = pd.cut(filtro_asa_norte["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "SQN 100": ["SQN 102", "SQN 103", "SQN 104", "SQN 105", "SQN 106", "SQN 107", "SQN 108", "SQN 109", "SQN 110", "SQN 111", "SQN 112", "SQN 113", "SQN 114", "SQN 115", "SQN 116"],
        "SQN 200": ["SQN 202", "SQN 203", "SQN 204", "SQN 205", "SQN 206", "SQN 207", "SQN 208", "SQN 209", "SQN 210", "SQN 211", "SQN 212", "SQN 213", "SQN 214", "SQN 215", "SQN 216"],
        "SQN 300": ["SQN 302", "SQN 303", "SQN 304", "SQN 305", "SQN 306", "SQN 307", "SQN 308", "SQN 309", "SQN 310", "SQN 311", "SQN 312", "SQN 313", "SQN 314", "SQN 315", "SQN 316"],
        "SQN 400": ["SQN 402", "SQN 403", "SQN 404", "SQN 405", "SQN 406", "SQN 407", "SQN 408", "SQN 409", "SQN 410", "SQN 411", "SQN 412", "SQN 413", "SQN 414", "SQN 415", "SQN 416"]
    }

    quadras_individuais = [
        "SQN 102", "SQN 103", "SQN 104", "SQN 105", "SQN 106", "SQN 107",
        "SQN 108", "SQN 109", "SQN 110", "SQN 111", "SQN 112", "SQN 113",
        "SQN 114", "SQN 115", "SQN 116", "SQN 202", "SQN 203", "SQN 204",
        "SQN 205", "SQN 206", "SQN 207", "SQN 208", "SQN 209", "SQN 210",
        "SQN 211", "SQN 212", "SQN 213", "SQN 214", "SQN 215", "SQN 216",
        "SQN 302", "SQN 303", "SQN 304", "SQN 305", "SQN 306", "SQN 307",
        "SQN 308", "SQN 309", "SQN 310", "SQN 311", "SQN 312", "SQN 313",
        "SQN 314", "SQN 315", "SQN 316", "SQN 402", "SQN 403", "SQN 404",
        "SQN 405", "SQN 406", "SQN 407", "SQN 408", "SQN 409", "SQN 410",
        "SQN 411", "SQN 412", "SQN 413", "SQN 414", "SQN 415", "SQN 416"
    ]

    for quadra in quadras_individuais:
        groups[quadra] = [quadra]

    # Calculate averages by cluster
    media_clusters = filtro_asa_norte.groupby("cluster")["valor_m2"].mean()

    # Calculate averages by group of area
    media_metragem = filtro_asa_norte.groupby("grupo_metragem")["valor_m2"].mean()

    # Calculate averages by number of rooms
    media_quartos = filtro_asa_norte.groupby("quartos_group")["valor_m2"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = filtro_asa_norte[filtro_asa_norte["quadra"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["valor_m2"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/ASA NORTE SEM VAGA.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)


    ################### ORGANIZA OS RESULTADOS ##################


    import openpyxl

    # Carregar os arquivos Excel originais
    input_file1 = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/ASA NORTE COM VAGA.xlsx"
    input_file2 = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/ASA NORTE SEM VAGA.xlsx"

    workbook1 = openpyxl.load_workbook(input_file1)
    workbook2 = openpyxl.load_workbook(input_file2)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/ASA NORTE.xlsx"
    result_workbook = openpyxl.Workbook()

    # Para cada aba no primeiro arquivo, copiar os resultados para o novo arquivo
    for sheet_name in workbook1.sheetnames:
        original_sheet = workbook1[sheet_name]
        result_sheet = result_workbook.create_sheet(title=sheet_name)

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

    # Para cada aba no segundo arquivo, copiar os resultados para o novo arquivo
    for sheet_name in workbook2.sheetnames:
        original_sheet = workbook2[sheet_name]
        result_sheet = result_workbook.create_sheet(title=sheet_name)

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

    # Salvar o novo arquivo Excel com os resultados finais
    result_workbook.save(output_file)
    print("Planilhas unidas e salvas em:", output_file)


    import os
    import openpyxl

    # Caminho do arquivo Excel com os resultados
    input_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/ASA NORTE.xlsx"

    # Carregar o arquivo Excel original
    workbook = openpyxl.load_workbook(input_file, data_only=True)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Organizados/ASA NORTE.xlsx"
    result_workbook = openpyxl.Workbook()

    # Remover a aba padrão criada automaticamente
    default_sheet = result_workbook.active
    result_workbook.remove(default_sheet)

    # Criar uma aba para os resultados finais
    result_sheet = result_workbook.create_sheet(title="ASA NORTE")

    # Para cada aba no arquivo original, copiar os resultados para a nova aba no arquivo final
    for sheet_name in workbook.sheetnames:
        original_sheet = workbook[sheet_name]

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

        result_sheet.append([])  # Adicionar duas linhas de espaço entre as tabelas

    # Salvar o novo arquivo Excel com os resultados finais no mesmo local do arquivo de origem
    result_workbook.save(output_file)
    print("Resultados finais salvos em:", output_file)

def funcao_asa_sul():

    ################### MEDIAS DO GRUPO COM VAGA ##################


    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Apartamento com bairro "ASA SUL COM VAGA"
    filtro_asa_sul = (df["bairro"] == "ASA SUL") & (df["tipo"] == "Apartamento") & (df["oferta"] == "Venda") & (df["vagas"].notnull())
    df_asa_sul = df[filtro_asa_sul]

    # Remover valores absurdos
    media_valor_m2 = df_asa_sul["valor_m2"].mean()
    limite_superior = media_valor_m2 * 2.0
    limite_inferior = media_valor_m2 * 0.5
    df_asa_sul = df_asa_sul[(df_asa_sul["valor_m2"] >= limite_inferior) & (df_asa_sul["valor_m2"] <= limite_superior)]

    # Clusterização
    X = df_asa_sul[["valor_m2"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    df_asa_sul["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = df_asa_sul.groupby("cluster")["valor_m2"].mean().sort_values().index
    df_asa_sul["cluster"] = df_asa_sul["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    df_asa_sul["quartos_group"] = pd.cut(df_asa_sul["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 50, 75, 90, 130, 160, 200, np.inf]
    labels = ["<50", "50-75", "75-90", "90-130", "130-160", "160-200", ">200"]
    df_asa_sul["grupo_metragem"] = pd.cut(df_asa_sul["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "SQS 100": ["SQS 102", "SQS 103", "SQS 104", "SQS 105", "SQS 106", "SQS 107", "SQS 108", "SQS 109", "SQS 110", "SQS 111", "SQS 112", "SQS 113", "SQS 114", "SQS 115", "SQS 116"],
        "SQS 200": ["SQS 202", "SQS 203", "SQS 204", "SQS 205", "SQS 206", "SQS 207", "SQS 208", "SQS 209", "SQS 210", "SQS 211", "SQS 212", "SQS 213", "SQS 214", "SQS 215", "SQS 216"],
        "SQS 300": ["SQS 302", "SQS 303", "SQS 304", "SQS 305", "SQS 306", "SQS 307", "SQS 308", "SQS 309", "SQS 310", "SQS 311", "SQS 312", "SQS 313", "SQS 314", "SQS 315", "SQS 316"],
        "SQS 400": ["SQS 402", "SQS 403", "SQS 404", "SQS 405", "SQS 406", "SQS 407", "SQS 408", "SQS 409", "SQS 410", "SQS 411", "SQS 412", "SQS 413", "SQS 414", "SQS 415", "SQS 416"],
        "SHIGS 700": ["SHIGS 702", "SHIGS 703", "SHIGS 704", "SHIGS 705", "SHIGS 706", "SHIGS 707", "SHIGS 708", "SHIGS 709", "SHIGS 710", "SHIGS 711", "SHIGS 712", "SHIGS 713", "SHIGS 714", "SHIGS 715", "SHIGS 716"]
    }

    quadras_individuais = [
        "SQS 102", "SQS 103", "SQS 104", "SQS 105", "SQS 106", "SQS 107",
        "SQS 108", "SQS 109", "SQS 110", "SQS 111", "SQS 112", "SQS 113",
        "SQS 114", "SQS 115", "SQS 116", "SQS 202", "SQS 203", "SQS 204",
        "SQS 205", "SQS 206", "SQS 207", "SQS 208", "SQS 209", "SQS 210",
        "SQS 211", "SQS 212", "SQS 213", "SQS 214", "SQS 215", "SQS 216",
        "SQS 302", "SQS 303", "SQS 304", "SQS 305", "SQS 306", "SQS 307",
        "SQS 308", "SQS 309", "SQS 310", "SQS 311", "SQS 312", "SQS 313",
        "SQS 314", "SQS 315", "SQS 316", "SQS 402", "SQS 403", "SQS 404",
        "SQS 405", "SQS 406", "SQS 407", "SQS 408", "SQS 409", "SQS 410",
        "SQS 411", "SQS 412", "SQS 413", "SQS 414", "SQS 415", "SQS 416",
        "SHIGS 702", "SHIGS 703", "SHIGS 704", "SHIGS 705", "SHIGS 706", 
        "SHIGS 707", "SHIGS 708", "SHIGS 709", "SHIGS 710", "SHIGS 711", 
        "SHIGS 712", "SHIGS 713", "SHIGS 714", "SHIGS 715", "SHIGS 716"
    ]

    for quadra in quadras_individuais:
        groups[quadra] = [quadra]

    # Calculate averages by cluster
    media_clusters = df_asa_sul.groupby("cluster")["valor_m2"].mean()

    # Calculate averages by group of area
    media_metragem = df_asa_sul.groupby("grupo_metragem")["valor_m2"].mean()

    # Calculate averages by number of rooms
    media_quartos = df_asa_sul.groupby("quartos_group")["valor_m2"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = df_asa_sul[df_asa_sul["quadra"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["valor_m2"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/ASA SUL COM VAGA.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)


    ################### MEDIAS DO GRUPO SEM VAGA ##################


    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Apartamento com bairro "ASA SUL SEM VAGA"
    filtro_asa_sul = (df["bairro"] == "ASA SUL") & (df["tipo"] == "Apartamento") & (df["oferta"] == "Venda") & (df["vagas"].isnull())
    df_asa_sul = df[filtro_asa_sul]

    # Remover valores absurdos
    media_valor_m2 = df_asa_sul["valor_m2"].mean()
    limite_superior = media_valor_m2 * 2.0
    limite_inferior = media_valor_m2 * 0.5
    df_asa_sul = df_asa_sul[(df_asa_sul["valor_m2"] >= limite_inferior) & (df_asa_sul["valor_m2"] <= limite_superior)]

    # Clusterização
    X = df_asa_sul[["valor_m2"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    df_asa_sul["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = df_asa_sul.groupby("cluster")["valor_m2"].mean().sort_values().index
    df_asa_sul["cluster"] = df_asa_sul["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    df_asa_sul["quartos_group"] = pd.cut(df_asa_sul["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 50, 75, 90, 130, 160, 200, np.inf]
    labels = ["<50", "50-75", "75-90", "90-130", "130-160", "160-200", ">200"]
    df_asa_sul["grupo_metragem"] = pd.cut(df_asa_sul["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "SQS 100": ["SQS 102", "SQS 103", "SQS 104", "SQS 105", "SQS 106", "SQS 107", "SQS 108", "SQS 109", "SQS 110", "SQS 111", "SQS 112", "SQS 113", "SQS 114", "SQS 115", "SQS 116"],
        "SQS 200": ["SQS 202", "SQS 203", "SQS 204", "SQS 205", "SQS 206", "SQS 207", "SQS 208", "SQS 209", "SQS 210", "SQS 211", "SQS 212", "SQS 213", "SQS 214", "SQS 215", "SQS 216"],
        "SQS 300": ["SQS 302", "SQS 303", "SQS 304", "SQS 305", "SQS 306", "SQS 307", "SQS 308", "SQS 309", "SQS 310", "SQS 311", "SQS 312", "SQS 313", "SQS 314", "SQS 315", "SQS 316"],
        "SQS 400": ["SQS 402", "SQS 403", "SQS 404", "SQS 405", "SQS 406", "SQS 407", "SQS 408", "SQS 409", "SQS 410", "SQS 411", "SQS 412", "SQS 413", "SQS 414", "SQS 415", "SQS 416"]
    }

    quadras_individuais = [
        "SQS 102", "SQS 103", "SQS 104", "SQS 105", "SQS 106", "SQS 107",
        "SQS 108", "SQS 109", "SQS 110", "SQS 111", "SQS 112", "SQS 113",
        "SQS 114", "SQS 115", "SQS 116", "SQS 202", "SQS 203", "SQS 204",
        "SQS 205", "SQS 206", "SQS 207", "SQS 208", "SQS 209", "SQS 210",
        "SQS 211", "SQS 212", "SQS 213", "SQS 214", "SQS 215", "SQS 216",
        "SQS 302", "SQS 303", "SQS 304", "SQS 305", "SQS 306", "SQS 307",
        "SQS 308", "SQS 309", "SQS 310", "SQS 311", "SQS 312", "SQS 313",
        "SQS 314", "SQS 315", "SQS 316", "SQS 402", "SQS 403", "SQS 404",
        "SQS 405", "SQS 406", "SQS 407", "SQS 408", "SQS 409", "SQS 410",
        "SQS 411", "SQS 412", "SQS 413", "SQS 414", "SQS 415", "SQS 416"
    ]

    for quadra in quadras_individuais:
        groups[quadra] = [quadra]

    # Calculate averages by cluster
    media_clusters = df_asa_sul.groupby("cluster")["valor_m2"].mean()

    # Calculate averages by group of area
    media_metragem = df_asa_sul.groupby("grupo_metragem")["valor_m2"].mean()

    # Calculate averages by number of rooms
    media_quartos = df_asa_sul.groupby("quartos_group")["valor_m2"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = df_asa_sul[df_asa_sul["quadra"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["valor_m2"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/ASA SUL SEM VAGA.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)


    ################### Organiza os rsultados ##################


    import openpyxl

    # Carregar os arquivos Excel originais
    input_file1 = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/ASA SUL COM VAGA.xlsx"
    input_file2 = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/ASA SUL SEM VAGA.xlsx"

    workbook1 = openpyxl.load_workbook(input_file1)
    workbook2 = openpyxl.load_workbook(input_file2)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/ASA SUL.xlsx"
    result_workbook = openpyxl.Workbook()

    # Para cada aba no primeiro arquivo, copiar os resultados para o novo arquivo
    for sheet_name in workbook1.sheetnames:
        original_sheet = workbook1[sheet_name]
        result_sheet = result_workbook.create_sheet(title=sheet_name)

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

    # Para cada aba no segundo arquivo, copiar os resultados para o novo arquivo
    for sheet_name in workbook2.sheetnames:
        original_sheet = workbook2[sheet_name]
        result_sheet = result_workbook.create_sheet(title=sheet_name)

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

    # Salvar o novo arquivo Excel com os resultados finais
    result_workbook.save(output_file)
    print("Planilhas unidas e salvas em:", output_file)


    import os
    import openpyxl

    # Caminho do arquivo Excel com os resultados
    input_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/ASA SUL.xlsx"

    # Carregar o arquivo Excel original
    workbook = openpyxl.load_workbook(input_file, data_only=True)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Organizados/ASA SUL.xlsx"
    result_workbook = openpyxl.Workbook()

    # Remover a aba padrão criada automaticamente
    default_sheet = result_workbook.active
    result_workbook.remove(default_sheet)

    # Criar uma aba para os resultados finais
    result_sheet = result_workbook.create_sheet(title="ASA SUL")

    # Para cada aba no arquivo original, copiar os resultados para a nova aba no arquivo final
    for sheet_name in workbook.sheetnames:
        original_sheet = workbook[sheet_name]

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

        result_sheet.append([])  # Adicionar duas linhas de espaço entre as tabelas

    # Salvar o novo arquivo Excel com os resultados finais no mesmo local do arquivo de origem
    result_workbook.save(output_file)
    print("Resultados finais salvos em:", output_file)

def funcao_guara_apt():
    
    ################### MEDIAS DO GRUPO COM VAGA ##################
    
    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Apartamento com bairro "GUARÁ APT COM VAGA"
    filtro_guara = ((df["cidade"] == "GUARA") | (df["cidade"] == "GUARÁ")) & (df["tipo"] == "Apartamento") & (df["oferta"] == "Venda")& (df["vagas"].notnull())
    filtro_guara = df[filtro_guara]

    # Remover valores absurdos
    media_valor_m2 = filtro_guara["valor_m2"].mean()
    limite_superior = media_valor_m2 * 2.0
    limite_inferior = media_valor_m2 * 0.5
    filtro_guara = filtro_guara[(filtro_guara["valor_m2"] >= limite_inferior) & (filtro_guara["valor_m2"] <= limite_superior)]

    # Clusterização
    X = filtro_guara[["valor_m2"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    filtro_guara["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = filtro_guara.groupby("cluster")["valor_m2"].mean().sort_values().index
    filtro_guara["cluster"] = filtro_guara["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    filtro_guara["quartos_group"] = pd.cut(filtro_guara["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 50, 75, 90, 130, 160, 200, np.inf]
    labels = ["<50", "50-75", "75-90", "90-130", "130-160", "160-200", ">200"]
    filtro_guara["grupo_metragem"] = pd.cut(filtro_guara["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "Guará 1": ["GUARA I"],
        "Guará 2": ["GUARA II"],
        "Lúcio Costa": ["QUADRAS ECONOMICAS LUCIO COSTA"],
        "Setor Industrial": ["SETOR INDUSTRIAL"]
    }

    # Calculate averages by cluster
    media_clusters = filtro_guara.groupby("cluster")["valor_m2"].mean()

    # Calculate averages by group of area
    media_metragem = filtro_guara.groupby("grupo_metragem")["valor_m2"].mean()

    # Calculate averages by number of rooms
    media_quartos = filtro_guara.groupby("quartos_group")["valor_m2"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = filtro_guara[filtro_guara["bairro"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["valor_m2"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/GUARA-APT COM VAGA.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)

    ################### MEDIAS DO GRUPO SEM VAGA ##################

    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Apartamento com bairro "GUARÁ APT COM VAGA"
    filtro_guara = ((df["cidade"] == "GUARA") | (df["cidade"] == "GUARÁ")) & (df["tipo"] == "Apartamento") & (df["oferta"] == "Venda")& (df["vagas"].isnull())
    filtro_guara = df[filtro_guara]

    # Remover valores absurdos
    media_valor_m2 = filtro_guara["valor_m2"].mean()
    limite_superior = media_valor_m2 * 2.0
    limite_inferior = media_valor_m2 * 0.5
    filtro_guara = filtro_guara[(filtro_guara["valor_m2"] >= limite_inferior) & (filtro_guara["valor_m2"] <= limite_superior)]

    # Clusterização
    X = filtro_guara[["valor_m2"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    filtro_guara["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = filtro_guara.groupby("cluster")["valor_m2"].mean().sort_values().index
    filtro_guara["cluster"] = filtro_guara["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    filtro_guara["quartos_group"] = pd.cut(filtro_guara["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 50, 75, 90, 130, 160, 200, np.inf]
    labels = ["<50", "50-75", "75-90", "90-130", "130-160", "160-200", ">200"]
    filtro_guara["grupo_metragem"] = pd.cut(filtro_guara["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "Guará 1": ["GUARA I"],
        "Guará 2": ["GUARA II"],
        "Lúcio Costa": ["QUADRAS ECONOMICAS LUCIO COSTA"],
        "Setor Industrial": ["SETOR INDUSTRIAL"]
    }

    # Calculate averages by cluster
    media_clusters = filtro_guara.groupby("cluster")["valor_m2"].mean()

    # Calculate averages by group of area
    media_metragem = filtro_guara.groupby("grupo_metragem")["valor_m2"].mean()

    # Calculate averages by number of rooms
    media_quartos = filtro_guara.groupby("quartos_group")["valor_m2"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = filtro_guara[filtro_guara["bairro"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["valor_m2"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/GUARA-APT SEM VAGA.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)

    ################### ORGANIZA OS DADOS ##################


    import openpyxl

    # Carregar os arquivos Excel originais
    input_file1 = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/GUARA-APT COM VAGA.xlsx"
    input_file2 = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/GUARA-APT SEM VAGA.xlsx"

    workbook1 = openpyxl.load_workbook(input_file1)
    workbook2 = openpyxl.load_workbook(input_file2)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/GUARA-APT.xlsx"
    result_workbook = openpyxl.Workbook()

    # Para cada aba no primeiro arquivo, copiar os resultados para o novo arquivo
    for sheet_name in workbook1.sheetnames:
        original_sheet = workbook1[sheet_name]
        result_sheet = result_workbook.create_sheet(title=sheet_name)

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

    # Para cada aba no segundo arquivo, copiar os resultados para o novo arquivo
    for sheet_name in workbook2.sheetnames:
        original_sheet = workbook2[sheet_name]
        result_sheet = result_workbook.create_sheet(title=sheet_name)

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

    # Salvar o novo arquivo Excel com os resultados finais
    result_workbook.save(output_file)
    print("Planilhas unidas e salvas em:", output_file)


    import os
    import openpyxl

    # Caminho do arquivo Excel com os resultados
    input_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/GUARA-APT.xlsx"

    # Carregar o arquivo Excel original
    workbook = openpyxl.load_workbook(input_file, data_only=True)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Organizados/GUARA-APT.xlsx"
    result_workbook = openpyxl.Workbook()

    # Remover a aba padrão criada automaticamente
    default_sheet = result_workbook.active
    result_workbook.remove(default_sheet)

    # Criar uma aba para os resultados finais
    result_sheet = result_workbook.create_sheet(title="GUARA-APT")

    # Para cada aba no arquivo original, copiar os resultados para a nova aba no arquivo final
    for sheet_name in workbook.sheetnames:
        original_sheet = workbook[sheet_name]

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

        result_sheet.append([])  # Adicionar duas linhas de espaço entre as tabelas

    # Salvar o novo arquivo Excel com os resultados finais no mesmo local do arquivo de origem
    result_workbook.save(output_file)
    print("Resultados finais salvos em:", output_file)

def funcao_guara_casa():
    
     ################### MEDIAS ##################
    
    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Casa com bairro "GUARÁ"
    filtro_guara = ((df["cidade"] == "GUARA") | (df["cidade"] == "GUARÁ")) & (df["tipo"] == "Casa") & (df["oferta"] == "Venda")
    filtro_guara = df[filtro_guara]

    # Remover valores absurdos
    media_valor_m2 = filtro_guara["valor_m2"].mean()
    limite_superior = media_valor_m2 * 2.0
    limite_inferior = media_valor_m2 * 0.5
    filtro_guara = filtro_guara[(filtro_guara["valor_m2"] >= limite_inferior) & (filtro_guara["valor_m2"] <= limite_superior)]

    # Clusterização
    X = filtro_guara[["valor_m2"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    filtro_guara["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = filtro_guara.groupby("cluster")["valor_m2"].mean().sort_values().index
    filtro_guara["cluster"] = filtro_guara["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    filtro_guara["quartos_group"] = pd.cut(filtro_guara["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 50, 75, 90, 130, 160, 200, np.inf]
    labels = ["<50", "50-75", "75-90", "90-130", "130-160", "160-200", ">200"]
    filtro_guara["grupo_metragem"] = pd.cut(filtro_guara["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "Guará 1": ["GUARA I"],
        "Guará 2": ["GUARA II"],
        "Lúcio Costa": ["QUADRAS ECONOMICAS LUCIO COSTA"],
        "Setor Industrial": ["SETOR INDUSTRIAL"]
    }

    # Calculate averages by cluster
    media_clusters = filtro_guara.groupby("cluster")["valor_m2"].mean()

    # Calculate averages by group of area
    media_metragem = filtro_guara.groupby("grupo_metragem")["valor_m2"].mean()

    # Calculate averages by number of rooms
    media_quartos = filtro_guara.groupby("quartos_group")["valor_m2"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = filtro_guara[filtro_guara["bairro"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["valor_m2"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/GUARA-CASA.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Resultados_Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)


    ################### ORGANIZA OS RESULTADOS ##################

    import os
    import openpyxl

    # Caminho do arquivo Excel com os resultados
    input_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/GUARA-CASA.xlsx"

    # Carregar o arquivo Excel original
    workbook = openpyxl.load_workbook(input_file, data_only=True)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Organizados/GUARA-CASA.xlsx"
    result_workbook = openpyxl.Workbook()

    # Remover a aba padrão criada automaticamente
    default_sheet = result_workbook.active
    result_workbook.remove(default_sheet)

    # Criar uma aba para os resultados finais
    result_sheet = result_workbook.create_sheet(title="GUARA-CASA")

    # Para cada aba no arquivo original, copiar os resultados para a nova aba no arquivo final
    for sheet_name in workbook.sheetnames:
        original_sheet = workbook[sheet_name]

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

        result_sheet.append([])  # Adicionar duas linhas de espaço entre as tabelas

    # Salvar o novo arquivo Excel com os resultados finais no mesmo local do arquivo de origem
    result_workbook.save(output_file)
    print("Resultados finais salvos em:", output_file)

def funcao_jardim_botanico():
    
     ################### MEDIAS ##################
    
    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Casa com bairro "JARDIM BOTANICO"
    filtro_jardim_botanico = (df["bairro"] == "JARDIM BOTANICO") & (df["tipo"] == "Casa") & (df["oferta"] == "Venda")
    df_jardim_botanico = df[filtro_jardim_botanico]

    # Remover valores absurdos
    media_preco = df_jardim_botanico["preco"].mean()
    limite_superior = media_preco * 2.0
    limite_inferior = media_preco * 0.5
    df_jardim_botanico = df_jardim_botanico[(df_jardim_botanico["preco"] >= limite_inferior) & (df_jardim_botanico["preco"] <= limite_superior)]

    # Clusterização
    X = df_jardim_botanico[["preco"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    df_jardim_botanico["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = df_jardim_botanico.groupby("cluster")["preco"].mean().sort_values().index
    df_jardim_botanico["cluster"] = df_jardim_botanico["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    df_jardim_botanico["quartos_group"] = pd.cut(df_jardim_botanico["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 100, 200, 300, 400, 500, np.inf]
    labels = ["<100", "100-200", "200-300", "300-400", "400-500", ">500"]
    df_jardim_botanico["grupo_metragem"] = pd.cut(df_jardim_botanico["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "JARDIM BOTANICO": [" "],
    }

    # Calculate averages by cluster
    media_clusters = df_jardim_botanico.groupby("cluster")["preco"].mean()

    # Calculate averages by group of area
    media_metragem = df_jardim_botanico.groupby("grupo_metragem")["preco"].mean()

    # Calculate averages by number of rooms
    media_quartos = df_jardim_botanico.groupby("quartos_group")["preco"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = df_jardim_botanico[df_jardim_botanico["quadra"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["preco"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/JARDIM BOTANICO.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Resultados_Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)

    ################### ORGANIZAR DADOS ##################

    import os
    import openpyxl

    # Caminho do arquivo Excel com os resultados
    input_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/JARDIM BOTANICO.xlsx"

    # Carregar o arquivo Excel original
    workbook = openpyxl.load_workbook(input_file, data_only=True)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Organizados/JARDIM BOTANICO.xlsx"
    result_workbook = openpyxl.Workbook()

    # Remover a aba padrão criada automaticamente
    default_sheet = result_workbook.active
    result_workbook.remove(default_sheet)

    # Criar uma aba para os resultados finais
    result_sheet = result_workbook.create_sheet(title="JARDIM BOTANICO")

    # Para cada aba no arquivo original, copiar os resultados para a nova aba no arquivo final
    for sheet_name in workbook.sheetnames:
        original_sheet = workbook[sheet_name]

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

        result_sheet.append([])  # Adicionar duas linhas de espaço entre as tabelas

    # Salvar o novo arquivo Excel com os resultados finais no mesmo local do arquivo de origem
    result_workbook.save(output_file)
    print("Resultados finais salvos em:", output_file)

def funcao_lago_norte():    
    ################### MEDIAS ##################

    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Casa com bairro "LAGO NORTE"
    filtro_lago_norte = (df["bairro"] == "LAGO NORTE") & (df["tipo"] == "Casa") & (df["oferta"] == "Venda") & (df["preco"] <3500000)
    df_lago_norte = df[filtro_lago_norte]

    # Remover valores absurdos
    media_preco = df_lago_norte["preco"].mean()
    limite_superior = media_preco * 2.0
    limite_inferior = media_preco * 0.5
    df_lago_norte = df_lago_norte[(df_lago_norte["preco"] >= limite_inferior) & (df_lago_norte["preco"] <= limite_superior)]

    # Clusterização
    X = df_lago_norte[["preco"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    df_lago_norte["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = df_lago_norte.groupby("cluster")["preco"].mean().sort_values().index
    df_lago_norte["cluster"] = df_lago_norte["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    df_lago_norte["quartos_group"] = pd.cut(df_lago_norte["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 400, 600, 800, 1000, np.inf]
    labels = ["<400", "400-600", "600-800", "800-1000", ">1000"]
    df_lago_norte["grupo_metragem"] = pd.cut(df_lago_norte["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "QIs": [
            "SHIN QI 1", "SHIN QI 2", "SHIN QI 3", "SHIN QI 4", "SHIN QI 5",
            "SHIN QI 6", "SHIN QI 7", "SHIN QI 8", "SHIN QI 9", "SHIN QI 10",
            "SHIN QI 11", "SHIN QI 12", "SHIN QI 13", "SHIN QI 14", "SHIN QI 15",
            "SHIN QI 16", "SHIN QI 17", "SHIN QI 18", "SHIN QI 19", "SHIN QI 20"
        ],
        "QLs": [
            "SHIN QL 1", "SHIN QL 2", "SHIN QL 3", "SHIN QL 4", "SHIN QL 5",
            "SHIN QL 6", "SHIN QL 7", "SHIN QL 8", "SHIN QL 9", "SHIN QL 10",
            "SHIN QL 11", "SHIN QL 12", "SHIN QL 13", "SHIN QL 14", "SHIN QL 15",
            "SHIN QL 16", "SHIN QL 17", "SHIN QL 18", "SHIN QL 19", "SHIN QL 20"
        ]
    }

    # Calculate averages by cluster
    media_clusters = df_lago_norte.groupby("cluster")["preco"].mean()

    # Calculate averages by group of area
    media_metragem = df_lago_norte.groupby("grupo_metragem")["preco"].mean()

    # Calculate averages by number of rooms
    media_quartos = df_lago_norte.groupby("quartos_group")["preco"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = df_lago_norte[df_lago_norte["quadra"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["preco"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/LAGO NORTE.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Resultados_Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)

    ################### ORGANIZAR RESULTADOS ##################

    import os
    import openpyxl

    # Caminho do arquivo Excel com os resultados
    input_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/LAGO NORTE.xlsx"

    # Carregar o arquivo Excel original
    workbook = openpyxl.load_workbook(input_file, data_only=True)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Organizados/LAGO NORTE.xlsx"
    result_workbook = openpyxl.Workbook()

    # Remover a aba padrão criada automaticamente
    default_sheet = result_workbook.active
    result_workbook.remove(default_sheet)

    # Criar uma aba para os resultados finais
    result_sheet = result_workbook.create_sheet(title="LAGO NORTE")

    # Para cada aba no arquivo original, copiar os resultados para a nova aba no arquivo final
    for sheet_name in workbook.sheetnames:
        original_sheet = workbook[sheet_name]

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

        result_sheet.append([])  # Adicionar duas linhas de espaço entre as tabelas

    # Salvar o novo arquivo Excel com os resultados finais no mesmo local do arquivo de origem
    result_workbook.save(output_file)
    print("Resultados finais salvos em:", output_file)

def funcao_lago_norte_luxo():

    ################### MEDIAS ##################

    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Casa com bairro "LAGO NORTE"
    filtro_lago_norte = (df["bairro"] == "LAGO NORTE") & (df["tipo"] == "Casa") & (df["oferta"] == "Venda") & (df["preco"] >3500000)
    df_lago_norte = df[filtro_lago_norte]

    # Remover valores absurdos
    media_preco = df_lago_norte["preco"].mean()
    limite_superior = media_preco * 2.0
    limite_inferior = media_preco * 0.5
    df_lago_norte = df_lago_norte[(df_lago_norte["preco"] >= limite_inferior) & (df_lago_norte["preco"] <= limite_superior)]

    # Clusterização
    X = df_lago_norte[["preco"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    df_lago_norte["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = df_lago_norte.groupby("cluster")["preco"].mean().sort_values().index
    df_lago_norte["cluster"] = df_lago_norte["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    df_lago_norte["quartos_group"] = pd.cut(df_lago_norte["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 400, 600, 800, 1000, np.inf]
    labels = ["<400", "400-600", "600-800", "800-1000", ">1000"]
    df_lago_norte["grupo_metragem"] = pd.cut(df_lago_norte["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "QIs": [
            "SHIN QI 1", "SHIN QI 2", "SHIN QI 3", "SHIN QI 4", "SHIN QI 5",
            "SHIN QI 6", "SHIN QI 7", "SHIN QI 8", "SHIN QI 9", "SHIN QI 10",
            "SHIN QI 11", "SHIN QI 12", "SHIN QI 13", "SHIN QI 14", "SHIN QI 15",
            "SHIN QI 16", "SHIN QI 17", "SHIN QI 18", "SHIN QI 19", "SHIN QI 20"
        ],
        "QLs": [
            "SHIN QL 1", "SHIN QL 2", "SHIN QL 3", "SHIN QL 4", "SHIN QL 5",
            "SHIN QL 6", "SHIN QL 7", "SHIN QL 8", "SHIN QL 9", "SHIN QL 10",
            "SHIN QL 11", "SHIN QL 12", "SHIN QL 13", "SHIN QL 14", "SHIN QL 15",
            "SHIN QL 16", "SHIN QL 17", "SHIN QL 18", "SHIN QL 19", "SHIN QL 20"
        ]
    }

    # Calculate averages by cluster
    media_clusters = df_lago_norte.groupby("cluster")["preco"].mean()

    # Calculate averages by group of area
    media_metragem = df_lago_norte.groupby("grupo_metragem")["preco"].mean()

    # Calculate averages by number of rooms
    media_quartos = df_lago_norte.groupby("quartos_group")["preco"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = df_lago_norte[df_lago_norte["quadra"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["preco"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/LAGO NORTE LUXO.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Resultados_Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)

    ################### ORGANIZAR RESULTADOS ##################
    import os
    import openpyxl

    # Caminho do arquivo Excel com os resultados
    input_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/LAGO NORTE LUXO.xlsx"

    # Carregar o arquivo Excel original
    workbook = openpyxl.load_workbook(input_file, data_only=True)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Organizados/LAGO NORTE LUXO.xlsx"
    result_workbook = openpyxl.Workbook()

    # Remover a aba padrão criada automaticamente
    default_sheet = result_workbook.active
    result_workbook.remove(default_sheet)

    # Criar uma aba para os resultados finais
    result_sheet = result_workbook.create_sheet(title="LAGO NORTE LUXO")

    # Para cada aba no arquivo original, copiar os resultados para a nova aba no arquivo final
    for sheet_name in workbook.sheetnames:
        original_sheet = workbook[sheet_name]

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

        result_sheet.append([])  # Adicionar duas linhas de espaço entre as tabelas

    # Salvar o novo arquivo Excel com os resultados finais no mesmo local do arquivo de origem
    result_workbook.save(output_file)
    print("Resultados finais salvos em:", output_file)

def funcao_lago_sul_luxo():

    ################### MEDIAS ##################

    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Casa com bairro "LAGO SUL"
    filtro_lago_sul = (df["bairro"] == "LAGO SUL") & (df["tipo"] == "Casa") & (df["oferta"] == "Venda")& (df["preco"] <5000000)
    df_lago_sul = df[filtro_lago_sul]

    # Remover valores absurdos
    media_preco = df_lago_sul["preco"].mean()
    limite_superior = media_preco * 2.0
    limite_inferior = media_preco * 0.5
    df_lago_sul = df_lago_sul[(df_lago_sul["preco"] >= limite_inferior) & (df_lago_sul["preco"] <= limite_superior)]

    # Clusterização
    X = df_lago_sul[["preco"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    df_lago_sul["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = df_lago_sul.groupby("cluster")["preco"].mean().sort_values().index
    df_lago_sul["cluster"] = df_lago_sul["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    df_lago_sul["quartos_group"] = pd.cut(df_lago_sul["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 400, 600, 800, 1000, np.inf]
    labels = ["<400", "400-600", "600-800", "800-1000", ">1000"]
    df_lago_sul["grupo_metragem"] = pd.cut(df_lago_sul["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "QIs": [
            "SHIS QI 1", "SHIS QI 2", "SHIS QI 3", "SHIS QI 4", "SHIS QI 5",
            "SHIS QI 6", "SHIS QI 7", "SHIS QI 8", "SHIS QI 9", "SHIS QI 10",
            "SHIS QI 11", "SHIS QI 12", "SHIS QI 13", "SHIS QI 14", "SHIS QI 15",
            "SHIS QI 16", "SHIS QI 17", "SHIS QI 18", "SHIS QI 19", "SHIS QI 20"
        ],
        "QLs": [
            "SHIS QL 1", "SHIS QL 2", "SHIS QL 3", "SHIS QL 4", "SHIS QL 5",
            "SHIS QL 6", "SHIS QL 7", "SHIS QL 8", "SHIS QL 9", "SHIS QL 10",
            "SHIS QL 11", "SHIS QL 12", "SHIS QL 13", "SHIS QL 14", "SHIS QL 15",
            "SHIS QL 16", "SHIS QL 17", "SHIS QL 18", "SHIS QL 19", "SHIS QL 20"
        ]
    }

    # Calculate averages by cluster
    media_clusters = df_lago_sul.groupby("cluster")["preco"].mean()

    # Calculate averages by group of area
    media_metragem = df_lago_sul.groupby("grupo_metragem")["preco"].mean()

    # Calculate averages by number of rooms
    media_quartos = df_lago_sul.groupby("quartos_group")["preco"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = df_lago_sul[df_lago_sul["quadra"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["preco"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/LAGO SUL LUXO.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Resultados_Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)


    ################### ORGANIZAR RESULTADOS ##################

    import os
    import openpyxl

    # Caminho do arquivo Excel com os resultados
    input_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/LAGO SUL LUXO.xlsx"

    # Carregar o arquivo Excel original
    workbook = openpyxl.load_workbook(input_file, data_only=True)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Organizados/LAGO SUL LUXO.xlsx"
    result_workbook = openpyxl.Workbook()

    # Remover a aba padrão criada automaticamente
    default_sheet = result_workbook.active
    result_workbook.remove(default_sheet)

    # Criar uma aba para os resultados finais
    result_sheet = result_workbook.create_sheet(title="LAGO SUL LUXO")

    # Para cada aba no arquivo original, copiar os resultados para a nova aba no arquivo final
    for sheet_name in workbook.sheetnames:
        original_sheet = workbook[sheet_name]

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

        result_sheet.append([])  # Adicionar duas linhas de espaço entre as tabelas

    # Salvar o novo arquivo Excel com os resultados finais no mesmo local do arquivo de origem
    result_workbook.save(output_file)
    print("Resultados finais salvos em:", output_file)

def funcao_lago_sul_alto_luxo():

    ################### MEDIAS ##################

    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Casa com bairro "LAGO SUL"
    filtro_lago_sul = (df["bairro"] == "LAGO SUL") & (df["tipo"] == "Casa") & (df["oferta"] == "Venda")& (df["preco"] >5000000)
    df_lago_sul = df[filtro_lago_sul]

    # Remover valores absurdos
    media_preco = df_lago_sul["preco"].mean()
    limite_superior = media_preco * 2.0
    limite_inferior = media_preco * 0.5
    df_lago_sul = df_lago_sul[(df_lago_sul["preco"] >= limite_inferior) & (df_lago_sul["preco"] <= limite_superior)]

    # Clusterização
    X = df_lago_sul[["preco"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    df_lago_sul["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = df_lago_sul.groupby("cluster")["preco"].mean().sort_values().index
    df_lago_sul["cluster"] = df_lago_sul["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    df_lago_sul["quartos_group"] = pd.cut(df_lago_sul["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 400, 600, 800, 1000, np.inf]
    labels = ["<400", "400-600", "600-800", "800-1000", ">1000"]
    df_lago_sul["grupo_metragem"] = pd.cut(df_lago_sul["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "QIs": [
            "SHIS QI 1", "SHIS QI 2", "SHIS QI 3", "SHIS QI 4", "SHIS QI 5",
            "SHIS QI 6", "SHIS QI 7", "SHIS QI 8", "SHIS QI 9", "SHIS QI 10",
            "SHIS QI 11", "SHIS QI 12", "SHIS QI 13", "SHIS QI 14", "SHIS QI 15",
            "SHIS QI 16", "SHIS QI 17", "SHIS QI 18", "SHIS QI 19", "SHIS QI 20"
        ],
        "QLs": [
            "SHIS QL 1", "SHIS QL 2", "SHIS QL 3", "SHIS QL 4", "SHIS QL 5",
            "SHIS QL 6", "SHIS QL 7", "SHIS QL 8", "SHIS QL 9", "SHIS QL 10",
            "SHIS QL 11", "SHIS QL 12", "SHIS QL 13", "SHIS QL 14", "SHIS QL 15",
            "SHIS QL 16", "SHIS QL 17", "SHIS QL 18", "SHIS QL 19", "SHIS QL 20"
        ]
    }

    # Calculate averages by cluster
    media_clusters = df_lago_sul.groupby("cluster")["preco"].mean()

    # Calculate averages by group of area
    media_metragem = df_lago_sul.groupby("grupo_metragem")["preco"].mean()

    # Calculate averages by number of rooms
    media_quartos = df_lago_sul.groupby("quartos_group")["preco"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = df_lago_sul[df_lago_sul["quadra"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["preco"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/LAGO SUL LUXO ALTO LUXO.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Resultados_Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)


    ################### ORGANIZAR RESULTADOS ##################

    import os
    import openpyxl

    # Caminho do arquivo Excel com os resultados
    input_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/LAGO SUL LUXO ALTO LUXO.xlsx"

    # Carregar o arquivo Excel original
    workbook = openpyxl.load_workbook(input_file, data_only=True)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Organizados/LAGO SUL ALTO LUXO.xlsx"
    result_workbook = openpyxl.Workbook()

    # Remover a aba padrão criada automaticamente
    default_sheet = result_workbook.active
    result_workbook.remove(default_sheet)

    # Criar uma aba para os resultados finais
    result_sheet = result_workbook.create_sheet(title="LAGO SUL ALTO LUXO")

    # Para cada aba no arquivo original, copiar os resultados para a nova aba no arquivo final
    for sheet_name in workbook.sheetnames:
        original_sheet = workbook[sheet_name]

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

        result_sheet.append([])  # Adicionar duas linhas de espaço entre as tabelas

    # Salvar o novo arquivo Excel com os resultados finais no mesmo local do arquivo de origem
    result_workbook.save(output_file)
    print("Resultados finais salvos em:", output_file)

def funcao_noroeste():
    
     ################### MEDIAS DO GRUPO COM VAGA ##################
    
    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Apartamento com bairro "NOROESTE"
    filtro_noroeste = (df["bairro"] == "NOROESTE") & (df["tipo"] == "Apartamento") & (df["oferta"] == "Venda") & (df["vagas"].notnull())
    df_noroeste = df[filtro_noroeste]

    # Remover valores absurdos
    media_valor_m2 = df_noroeste["valor_m2"].mean()
    limite_superior = media_valor_m2 * 2.0
    limite_inferior = media_valor_m2 * 0.5
    df_noroeste = df_noroeste[(df_noroeste["valor_m2"] >= limite_inferior) & (df_noroeste["valor_m2"] <= limite_superior)]

    # Clusterização
    X = df_noroeste[["valor_m2"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    df_noroeste["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = df_noroeste.groupby("cluster")["valor_m2"].mean().sort_values().index
    df_noroeste["cluster"] = df_noroeste["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    df_noroeste["quartos_group"] = pd.cut(df_noroeste["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 50, 75, 90, 130, 160, 200, np.inf]
    labels = ["<50", "50-75", "75-90", "90-130", "130-160", "160-200", ">200"]
    df_noroeste["grupo_metragem"] = pd.cut(df_noroeste["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "CLNW": ["CLNW 10/11", "CLNW 2/3", "CLNW 4/5", "CLNW 6/7"],
        "SQNW 100": ["SQNW 106", "SQNW 107", "SQNW 108", "SQNW 110"],
        "SQNW 300": ["SQNW 309", "SQNW 310", "SQNW 311"]
    }

    # Calculate averages by cluster
    media_clusters = df_noroeste.groupby("cluster")["valor_m2"].mean()

    # Calculate averages by group of area
    media_metragem = df_noroeste.groupby("grupo_metragem")["valor_m2"].mean()

    # Calculate averages by number of rooms
    media_quartos = df_noroeste.groupby("quartos_group")["valor_m2"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = df_noroeste[df_noroeste["quadra"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["valor_m2"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/NOROESTE COM VAGA.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)


    ################### MEDIAS DO GRUPO SEM VAGA ##################

    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Apartamento com bairro "NOROESTE"
    filtro_noroeste = (df["bairro"] == "NOROESTE") & (df["tipo"] == "Apartamento") & (df["oferta"] == "Venda") & (df["vagas"].isnull())
    df_noroeste = df[filtro_noroeste]

    # Remover valores absurdos
    media_valor_m2 = df_noroeste["valor_m2"].mean()
    limite_superior = media_valor_m2 * 2.0
    limite_inferior = media_valor_m2 * 0.5
    df_noroeste = df_noroeste[(df_noroeste["valor_m2"] >= limite_inferior) & (df_noroeste["valor_m2"] <= limite_superior)]

    # Clusterização
    X = df_noroeste[["valor_m2"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    df_noroeste["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = df_noroeste.groupby("cluster")["valor_m2"].mean().sort_values().index
    df_noroeste["cluster"] = df_noroeste["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    df_noroeste["quartos_group"] = pd.cut(df_noroeste["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 50, 75, 90, 130, 160, 200, np.inf]
    labels = ["<50", "50-75", "75-90", "90-130", "130-160", "160-200", ">200"]
    df_noroeste["grupo_metragem"] = pd.cut(df_noroeste["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "CLNW": ["CLNW 10/11", "CLNW 2/3", "CLNW 4/5", "CLNW 6/7"],
        "SQNW 100": ["SQNW 106", "SQNW 107", "SQNW 108", "SQNW 110"],
        "SQNW 300": ["SQNW 309", "SQNW 310", "SQNW 311"]
    }

    # Calculate averages by cluster
    media_clusters = df_noroeste.groupby("cluster")["valor_m2"].mean()

    # Calculate averages by group of area
    media_metragem = df_noroeste.groupby("grupo_metragem")["valor_m2"].mean()

    # Calculate averages by number of rooms
    media_quartos = df_noroeste.groupby("quartos_group")["valor_m2"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = df_noroeste[df_noroeste["quadra"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["valor_m2"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/NOROESTE SEM VAGA.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)

    ################### ORGANIZA OS RESULTADOS ##################

    import openpyxl

    # Carregar os arquivos Excel originais
    input_file1 = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/NOROESTE COM VAGA.xlsx"
    input_file2 = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/NOROESTE SEM VAGA.xlsx"

    workbook1 = openpyxl.load_workbook(input_file1)
    workbook2 = openpyxl.load_workbook(input_file2)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/NOROESTE.xlsx"
    result_workbook = openpyxl.Workbook()

    # Para cada aba no primeiro arquivo, copiar os resultados para o novo arquivo
    for sheet_name in workbook1.sheetnames:
        original_sheet = workbook1[sheet_name]
        result_sheet = result_workbook.create_sheet(title=sheet_name)

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

    # Para cada aba no segundo arquivo, copiar os resultados para o novo arquivo
    for sheet_name in workbook2.sheetnames:
        original_sheet = workbook2[sheet_name]
        result_sheet = result_workbook.create_sheet(title=sheet_name)

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

    # Salvar o novo arquivo Excel com os resultados finais
    result_workbook.save(output_file)
    print("Planilhas unidas e salvas em:", output_file)


    import os
    import openpyxl

    # Caminho do arquivo Excel com os resultados
    input_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/NOROESTE.xlsx"

    # Carregar o arquivo Excel original
    workbook = openpyxl.load_workbook(input_file, data_only=True)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Organizados/NOROESTE.xlsx"
    result_workbook = openpyxl.Workbook()

    # Remover a aba padrão criada automaticamente
    default_sheet = result_workbook.active
    result_workbook.remove(default_sheet)

    # Criar uma aba para os resultados finais
    result_sheet = result_workbook.create_sheet(title="NOROESTE")

    # Para cada aba no arquivo original, copiar os resultados para a nova aba no arquivo final
    for sheet_name in workbook.sheetnames:
        original_sheet = workbook[sheet_name]

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

        result_sheet.append([])  # Adicionar duas linhas de espaço entre as tabelas

    # Salvar o novo arquivo Excel com os resultados finais no mesmo local do arquivo de origem
    result_workbook.save(output_file)
    print("Resultados finais salvos em:", output_file)

def funcao_park_sul():
    
     ################### MEDIAS DO GRUPO COM VAGA ##################
    
    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Apartamento com bairro "PARK SUL COM VAGA"
    filtro_park_sul = (df["bairro"] == "PARK SUL") & (df["tipo"] == "Apartamento") & (df["oferta"] == "Venda") & (df["vagas"].notnull())
    df_park_sul = df[filtro_park_sul]

    # Remover valores absurdos
    media_valor_m2 = df_park_sul["valor_m2"].mean()
    limite_superior = media_valor_m2 * 2.0
    limite_inferior = media_valor_m2 * 0.5
    df_park_sul = df_park_sul[(df_park_sul["valor_m2"] >= limite_inferior) & (df_park_sul["valor_m2"] <= limite_superior)]

    # Clusterização
    X = df_park_sul[["valor_m2"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    df_park_sul["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = df_park_sul.groupby("cluster")["valor_m2"].mean().sort_values().index
    df_park_sul["cluster"] = df_park_sul["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    df_park_sul["quartos_group"] = pd.cut(df_park_sul["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 50, 75, 90, 130, 160, 200, np.inf]
    labels = ["<50", "50-75", "75-90", "90-130", "130-160", "160-200", ">200"]
    df_park_sul["grupo_metragem"] = pd.cut(df_park_sul["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "PARK SUL": [" "],
    }

    # Calculate averages by cluster
    media_clusters = df_park_sul.groupby("cluster")["valor_m2"].mean()

    # Calculate averages by group of area
    media_metragem = df_park_sul.groupby("grupo_metragem")["valor_m2"].mean()

    # Calculate averages by number of rooms
    media_quartos = df_park_sul.groupby("quartos_group")["valor_m2"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = df_park_sul[df_park_sul["quadra"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["valor_m2"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/PARK SUL COM VAGA.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)

    ################### MEDIAS DO GRUPO SEM VAGA ##################

    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Apartamento com bairro "PARK SUL SEM VAGA"
    filtro_park_sul = (df["bairro"] == "PARK SUL") & (df["tipo"] == "Apartamento") & (df["oferta"] == "Venda") & (df["vagas"].isnull())
    df_park_sul = df[filtro_park_sul]

    # Remover valores absurdos
    media_valor_m2 = df_park_sul["valor_m2"].mean()
    limite_superior = media_valor_m2 * 2.0
    limite_inferior = media_valor_m2 * 0.5
    df_park_sul = df_park_sul[(df_park_sul["valor_m2"] >= limite_inferior) & (df_park_sul["valor_m2"] <= limite_superior)]

    # Clusterização
    X = df_park_sul[["valor_m2"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    df_park_sul["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = df_park_sul.groupby("cluster")["valor_m2"].mean().sort_values().index
    df_park_sul["cluster"] = df_park_sul["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    df_park_sul["quartos_group"] = pd.cut(df_park_sul["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 50, 75, 90, 130, 160, 200, np.inf]
    labels = ["<50", "50-75", "75-90", "90-130", "130-160", "160-200", ">200"]
    df_park_sul["grupo_metragem"] = pd.cut(df_park_sul["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "PARK SUL": [" "],
    }

    # Calculate averages by cluster
    media_clusters = df_park_sul.groupby("cluster")["valor_m2"].mean()

    # Calculate averages by group of area
    media_metragem = df_park_sul.groupby("grupo_metragem")["valor_m2"].mean()

    # Calculate averages by number of rooms
    media_quartos = df_park_sul.groupby("quartos_group")["valor_m2"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = df_park_sul[df_park_sul["quadra"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["valor_m2"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/PARK SUL SEM VAGA.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)

    ################### ORGANIZAR RESULTADOS ##################

    import openpyxl

    # Carregar os arquivos Excel originais
    input_file1 = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/PARK SUL COM VAGA.xlsx"
    input_file2 = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/PARK SUL SEM VAGA.xlsx"

    workbook1 = openpyxl.load_workbook(input_file1)
    workbook2 = openpyxl.load_workbook(input_file2)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/PARK SUL.xlsx"
    result_workbook = openpyxl.Workbook()

    # Para cada aba no primeiro arquivo, copiar os resultados para o novo arquivo
    for sheet_name in workbook1.sheetnames:
        original_sheet = workbook1[sheet_name]
        result_sheet = result_workbook.create_sheet(title=sheet_name)

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

    # Para cada aba no segundo arquivo, copiar os resultados para o novo arquivo
    for sheet_name in workbook2.sheetnames:
        original_sheet = workbook2[sheet_name]
        result_sheet = result_workbook.create_sheet(title=sheet_name)

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

    # Salvar o novo arquivo Excel com os resultados finais
    result_workbook.save(output_file)
    print("Planilhas unidas e salvas em:", output_file)


    import os
    import openpyxl

    # Caminho do arquivo Excel com os resultados
    input_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/PARK SUL.xlsx"

    # Carregar o arquivo Excel original
    workbook = openpyxl.load_workbook(input_file, data_only=True)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Organizados/PARK SUL.xlsx"
    result_workbook = openpyxl.Workbook()

    # Remover a aba padrão criada automaticamente
    default_sheet = result_workbook.active
    result_workbook.remove(default_sheet)

    # Criar uma aba para os resultados finais
    result_sheet = result_workbook.create_sheet(title="PARK SUL")

    # Para cada aba no arquivo original, copiar os resultados para a nova aba no arquivo final
    for sheet_name in workbook.sheetnames:
        original_sheet = workbook[sheet_name]

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

        result_sheet.append([])  # Adicionar duas linhas de espaço entre as tabelas

    # Salvar o novo arquivo Excel com os resultados finais no mesmo local do arquivo de origem
    result_workbook.save(output_file)
    print("Resultados finais salvos em:", output_file)

def funcao_plano_piloto():
    
     ################### MEDIAS DO GRUPO COM VAGA ##################
    
    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Apartamento com bairro "PLANO PILOTO COM VAGA"
    filtro_plano_piloto = ((df["bairro"] == "SUDOESTE") | (df["bairro"] == "NOROESTE")| (df["bairro"] == "ASA SUL")| (df["bairro"] == "ASA NORTE")) & (df["tipo"] == "Apartamento") & (df["oferta"] == "Venda") & (df["vagas"].notnull())
    filtro_plano_piloto = df[filtro_plano_piloto]

    # Remover valores absurdos
    media_valor_m2 = filtro_plano_piloto["valor_m2"].mean()
    limite_superior = media_valor_m2 * 2.0
    limite_inferior = media_valor_m2 * 0.5
    filtro_plano_piloto = filtro_plano_piloto[(filtro_plano_piloto["valor_m2"] >= limite_inferior) & (filtro_plano_piloto["valor_m2"] <= limite_superior)]

    # Clusterização
    X = filtro_plano_piloto[["valor_m2"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    filtro_plano_piloto["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = filtro_plano_piloto.groupby("cluster")["valor_m2"].mean().sort_values().index
    filtro_plano_piloto["cluster"] = filtro_plano_piloto["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    filtro_plano_piloto["quartos_group"] = pd.cut(filtro_plano_piloto["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 50, 75, 90, 130, 160, 200, np.inf]
    labels = ["<50", "50-75", "75-90", "90-130", "130-160", "160-200", ">200"]
    filtro_plano_piloto["grupo_metragem"] = pd.cut(filtro_plano_piloto["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "Sudoeste": ["SUDOESTE"],
        "Noroeste": ["NOROESTE"],
        "Asa Sul": ["ASA SUL"],
        "Asa Norte": ["ASA NORTE"]
    }

    # Calculate averages by cluster
    media_clusters = filtro_plano_piloto.groupby("cluster")["valor_m2"].mean()

    # Calculate averages by group of area
    media_metragem = filtro_plano_piloto.groupby("grupo_metragem")["valor_m2"].mean()

    # Calculate averages by number of rooms
    media_quartos = filtro_plano_piloto.groupby("quartos_group")["valor_m2"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = filtro_plano_piloto[filtro_plano_piloto["bairro"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["valor_m2"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/PLANO PILOTO COM VAGA.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)

    ################### MEDIAS DO GRUPO SEM VAGA ##################

    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Apartamento com bairro "PLANO PILOTO SEM VAGA"
    filtro_plano_piloto = ((df["bairro"] == "SUDOESTE") | (df["bairro"] == "NOROESTE")| (df["bairro"] == "ASA SUL")| (df["bairro"] == "ASA NORTE")) & (df["tipo"] == "Apartamento") & (df["oferta"] == "Venda") & (df["vagas"].isnull())
    filtro_plano_piloto = df[filtro_plano_piloto]

    # Remover valores absurdos
    media_valor_m2 = filtro_plano_piloto["valor_m2"].mean()
    limite_superior = media_valor_m2 * 2.0
    limite_inferior = media_valor_m2 * 0.5
    filtro_plano_piloto = filtro_plano_piloto[(filtro_plano_piloto["valor_m2"] >= limite_inferior) & (filtro_plano_piloto["valor_m2"] <= limite_superior)]

    # Clusterização
    X = filtro_plano_piloto[["valor_m2"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    filtro_plano_piloto["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = filtro_plano_piloto.groupby("cluster")["valor_m2"].mean().sort_values().index
    filtro_plano_piloto["cluster"] = filtro_plano_piloto["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    filtro_plano_piloto["quartos_group"] = pd.cut(filtro_plano_piloto["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 50, 75, 90, 130, 160, 200, np.inf]
    labels = ["<50", "50-75", "75-90", "90-130", "130-160", "160-200", ">200"]
    filtro_plano_piloto["grupo_metragem"] = pd.cut(filtro_plano_piloto["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "Sudoeste": ["SUDOESTE"],
        "Noroeste": ["NOROESTE"],
        "Asa Sul": ["ASA SUL"],
        "Asa Norte": ["ASA NORTE"]
    }

    # Calculate averages by cluster
    media_clusters = filtro_plano_piloto.groupby("cluster")["valor_m2"].mean()

    # Calculate averages by group of area
    media_metragem = filtro_plano_piloto.groupby("grupo_metragem")["valor_m2"].mean()

    # Calculate averages by number of rooms
    media_quartos = filtro_plano_piloto.groupby("quartos_group")["valor_m2"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = filtro_plano_piloto[filtro_plano_piloto["bairro"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["valor_m2"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/PLANO PILOTO SEM VAGA.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)

    ################### ORGANIZAR RESULTADOS ##################


    import openpyxl

    # Carregar os arquivos Excel originais
    input_file1 = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/PLANO PILOTO COM VAGA.xlsx"
    input_file2 = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/PLANO PILOTO SEM VAGA.xlsx"

    workbook1 = openpyxl.load_workbook(input_file1)
    workbook2 = openpyxl.load_workbook(input_file2)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/PLANO PILOTO.xlsx"
    result_workbook = openpyxl.Workbook()

    # Para cada aba no primeiro arquivo, copiar os resultados para o novo arquivo
    for sheet_name in workbook1.sheetnames:
        original_sheet = workbook1[sheet_name]
        result_sheet = result_workbook.create_sheet(title=sheet_name)

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

    # Para cada aba no segundo arquivo, copiar os resultados para o novo arquivo
    for sheet_name in workbook2.sheetnames:
        original_sheet = workbook2[sheet_name]
        result_sheet = result_workbook.create_sheet(title=sheet_name)

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

    # Salvar o novo arquivo Excel com os resultados finais
    result_workbook.save(output_file)
    print("Planilhas unidas e salvas em:", output_file)


    import os
    import openpyxl

    # Caminho do arquivo Excel com os resultados
    input_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/PLANO PILOTO.xlsx"

    # Carregar o arquivo Excel original
    workbook = openpyxl.load_workbook(input_file, data_only=True)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Organizados/PLANO PILOTO.xlsx"
    result_workbook = openpyxl.Workbook()

    # Remover a aba padrão criada automaticamente
    default_sheet = result_workbook.active
    result_workbook.remove(default_sheet)

    # Criar uma aba para os resultados finais
    result_sheet = result_workbook.create_sheet(title="PLANO PILOTO")

    # Para cada aba no arquivo original, copiar os resultados para a nova aba no arquivo final
    for sheet_name in workbook.sheetnames:
        original_sheet = workbook[sheet_name]

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

        result_sheet.append([])  # Adicionar duas linhas de espaço entre as tabelas

    # Salvar o novo arquivo Excel com os resultados finais no mesmo local do arquivo de origem
    result_workbook.save(output_file)
    print("Resultados finais salvos em:", output_file)

def funcao_sudoeste():
    
     ################### MEDIAS DO GRUPO COM VAGA ##################
    
    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Apartamento com bairro "SUDOESTE COM VAGA"
    filtro_sudoeste = (df["bairro"] == "SUDOESTE") & (df["tipo"] == "Apartamento") & (df["oferta"] == "Venda") & (df["vagas"].notnull())
    df_sudoeste = df[filtro_sudoeste]

    # Remover valores absurdos
    media_valor_m2 = df_sudoeste["valor_m2"].mean()
    limite_superior = media_valor_m2 * 2.0
    limite_inferior = media_valor_m2 * 0.5
    df_sudoeste = df_sudoeste[(df_sudoeste["valor_m2"] >= limite_inferior) & (df_sudoeste["valor_m2"] <= limite_superior)]

    # Clusterização
    X = df_sudoeste[["valor_m2"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    df_sudoeste["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = df_sudoeste.groupby("cluster")["valor_m2"].mean().sort_values().index
    df_sudoeste["cluster"] = df_sudoeste["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    df_sudoeste["quartos_group"] = pd.cut(df_sudoeste["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 50, 75, 90, 130, 160, 200, np.inf]
    labels = ["<50", "50-75", "75-90", "90-130", "130-160", "160-200", ">200"]
    df_sudoeste["grupo_metragem"] = pd.cut(df_sudoeste["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "CCSW": ["CCSW 1", "CCSW 2", "CCSW 3"],
        "CLSW": ["CLSW 101", "CLSW 103", "CLSW 105", "CLSW 301", "CLSW 302", "CLSW 303", "CLSW 304"],
        "QMSW": ["QMSW 4", "QMSW 5", "QMSW 6"],
        "SQSW (SEM SQSW 300)": ["SQSW 100", "SQSW 101", "SQSW 102", "SQSW 103", "SQSW 104", "SQSW 105", "SQSW 504"],
        "SQSW 300": ["SQSW 300", "SQSW 301", "SQSW 302", "SQSW 303", "SQSW 304", "SQSW 305", "SQSW 306"],
        "QMSW": ["QMSW 5", "QMSW 6"]
    }


    # Calculate averages by cluster
    media_clusters = df_sudoeste.groupby("cluster")["valor_m2"].mean()

    # Calculate averages by group of area
    media_metragem = df_sudoeste.groupby("grupo_metragem")["valor_m2"].mean()

    # Calculate averages by number of rooms
    media_quartos = df_sudoeste.groupby("quartos_group")["valor_m2"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = df_sudoeste[df_sudoeste["quadra"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["valor_m2"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/SUDOESTE COM VAGA.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)

     ################### MEDIAS DO GRUPO SEM VAGA ##################

    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Apartamento com bairro "SUDOESTE SEM VAGA"
    filtro_sudoeste = (df["bairro"] == "SUDOESTE") & (df["tipo"] == "Apartamento") & (df["oferta"] == "Venda") & (df["vagas"].isnull())
    df_sudoeste = df[filtro_sudoeste]

    # Remover valores absurdos
    media_valor_m2 = df_sudoeste["valor_m2"].mean()
    limite_superior = media_valor_m2 * 2.0
    limite_inferior = media_valor_m2 * 0.5
    df_sudoeste = df_sudoeste[(df_sudoeste["valor_m2"] >= limite_inferior) & (df_sudoeste["valor_m2"] <= limite_superior)]

    # Clusterização
    X = df_sudoeste[["valor_m2"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    df_sudoeste["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = df_sudoeste.groupby("cluster")["valor_m2"].mean().sort_values().index
    df_sudoeste["cluster"] = df_sudoeste["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    df_sudoeste["quartos_group"] = pd.cut(df_sudoeste["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 50, 75, 90, 130, 160, 200, np.inf]
    labels = ["<50", "50-75", "75-90", "90-130", "130-160", "160-200", ">200"]
    df_sudoeste["grupo_metragem"] = pd.cut(df_sudoeste["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "CCSW": ["CCSW 1", "CCSW 2", "CCSW 3"],
        "CLSW": ["CLSW 101", "CLSW 103", "CLSW 105", "CLSW 301", "CLSW 302", "CLSW 303", "CLSW 304"],
        "QMSW": ["QMSW 4", "QMSW 5", "QMSW 6"],
        "SQSW (SEM SQSW 300)": ["SQSW 100", "SQSW 101", "SQSW 102", "SQSW 103", "SQSW 104", "SQSW 105", "SQSW 504"],
        "SQSW 300": ["SQSW 300", "SQSW 301", "SQSW 302", "SQSW 303", "SQSW 304", "SQSW 305", "SQSW 306"],
        "QMSW": ["QMSW 5", "QMSW 6"]
    }


    # Calculate averages by cluster
    media_clusters = df_sudoeste.groupby("cluster")["valor_m2"].mean()

    # Calculate averages by group of area
    media_metragem = df_sudoeste.groupby("grupo_metragem")["valor_m2"].mean()

    # Calculate averages by number of rooms
    media_quartos = df_sudoeste.groupby("quartos_group")["valor_m2"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = df_sudoeste[df_sudoeste["quadra"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["valor_m2"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/SUDOESTE SEM VAGA.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)

    ################### ORGANIZA RESULTADOS ##################


    import openpyxl

    # Carregar os arquivos Excel originais
    input_file1 = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/SUDOESTE COM VAGA.xlsx"
    input_file2 = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/SUDOESTE SEM VAGA.xlsx"

    workbook1 = openpyxl.load_workbook(input_file1)
    workbook2 = openpyxl.load_workbook(input_file2)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/SUDOESTE.xlsx"
    result_workbook = openpyxl.Workbook()

    # Para cada aba no primeiro arquivo, copiar os resultados para o novo arquivo
    for sheet_name in workbook1.sheetnames:
        original_sheet = workbook1[sheet_name]
        result_sheet = result_workbook.create_sheet(title=sheet_name)

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

    # Para cada aba no segundo arquivo, copiar os resultados para o novo arquivo
    for sheet_name in workbook2.sheetnames:
        original_sheet = workbook2[sheet_name]
        result_sheet = result_workbook.create_sheet(title=sheet_name)

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

    # Salvar o novo arquivo Excel com os resultados finais
    result_workbook.save(output_file)
    print("Planilhas unidas e salvas em:", output_file)


    import os
    import openpyxl

    # Caminho do arquivo Excel com os resultados
    input_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/SUDOESTE.xlsx"

    # Carregar o arquivo Excel original
    workbook = openpyxl.load_workbook(input_file, data_only=True)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Organizados/SUDOESTE.xlsx"
    result_workbook = openpyxl.Workbook()

    # Remover a aba padrão criada automaticamente
    default_sheet = result_workbook.active
    result_workbook.remove(default_sheet)

    # Criar uma aba para os resultados finais
    result_sheet = result_workbook.create_sheet(title="SUDOESTE")

    # Para cada aba no arquivo original, copiar os resultados para a nova aba no arquivo final
    for sheet_name in workbook.sheetnames:
        original_sheet = workbook[sheet_name]

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

        result_sheet.append([])  # Adicionar duas linhas de espaço entre as tabelas

    # Salvar o novo arquivo Excel com os resultados finais no mesmo local do arquivo de origem
    result_workbook.save(output_file)
    print("Resultados finais salvos em:", output_file)

def funcao_vicente_pires():
    
     ################### MEDIAS ##################
    
    import pandas as pd
    from sklearn.cluster import KMeans
    import numpy as np


    # Carregar o arquivo CSV
    input_file = "/Users/imac/Desktop/ColetasLimpa - Coleta1.csv"
    df = pd.read_csv(input_file, sep=",", thousands=".", decimal=",")

    # Filtrar somente os registros de Casa com bairro "VICENTE PIRES"
    filtro_vicente_pires = (df["bairro"] == "VICENTE PIRES") & (df["tipo"] == "Casa") & (df["oferta"] == "Venda")
    df_vicente_pires = df[filtro_vicente_pires]

    # Remover valores absurdos
    media_valor_m2 = df_vicente_pires["valor_m2"].mean()
    limite_superior = media_valor_m2 * 2.0
    limite_inferior = media_valor_m2 * 0.5
    df_vicente_pires = df_vicente_pires[(df_vicente_pires["valor_m2"] >= limite_inferior) & (df_vicente_pires["valor_m2"] <= limite_superior)]

    # Clusterização
    X = df_vicente_pires[["valor_m2"]].values
    kmeans = KMeans(n_clusters=9, random_state=42)
    df_vicente_pires["cluster"] = kmeans.fit_predict(X)

    # Organizar clusters de forma crescente
    cluster_order = df_vicente_pires.groupby("cluster")["valor_m2"].mean().sort_values().index
    df_vicente_pires["cluster"] = df_vicente_pires["cluster"].replace(dict(zip(cluster_order, range(len(cluster_order)))))

    # Definir os Números de quartos
    quartos_bins = [0, 1, 2, 3, 4, np.inf]
    quartos_labels = ["1", "2", "3", "4", "+5"]
    df_vicente_pires["quartos_group"] = pd.cut(df_vicente_pires["quartos"], bins=quartos_bins, labels=quartos_labels)

    # Definir os grupos de metragem
    bins = [0, 100, 200, 300, 400, 600, 800, np.inf]
    labels = ["<100", "100-200", "200-300", "300-400", "400-600", "600-800", ">800"]
    df_vicente_pires["grupo_metragem"] = pd.cut(df_vicente_pires["area_util"], bins=bins, labels=labels)

    # Definir os grupos de quadras
    groups = {
        "VICENTE PIRES": [" "],
    }

    # Calculate averages by cluster
    media_clusters = df_vicente_pires.groupby("cluster")["valor_m2"].mean()

    # Calculate averages by group of area
    media_metragem = df_vicente_pires.groupby("grupo_metragem")["valor_m2"].mean()

    # Calculate averages by number of rooms
    media_quartos = df_vicente_pires.groupby("quartos_group")["valor_m2"].mean()

    # Calculate averages by block
    media_quadras = {}
    for group, blocks in groups.items():
        filtered_df = df_vicente_pires[df_vicente_pires["quadra"].isin(blocks)]
        media_cluster = filtered_df.groupby("cluster")["valor_m2"].mean()
        media_quadras[group] = media_cluster

    # Save results to an Excel workbook
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/VICENTE PIRES.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for sheet_name, data in {
            "Resultados_Clusters": media_clusters,
            "Metragem": media_metragem,
            "Quartos": media_quartos,
            **media_quadras
        }.items():
            data.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)
            
            worksheet = writer.sheets[sheet_name]
            worksheet.write(0, 0, sheet_name)

    print("Processo concluído. Resultados salvos em:", output_file)

    ################### ORGANIZA RESULTADOS ##################

    import os
    import openpyxl

    # Caminho do arquivo Excel com os resultados
    input_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Sujos/VICENTE PIRES.xlsx"

    # Carregar o arquivo Excel original
    workbook = openpyxl.load_workbook(input_file, data_only=True)

    # Criar um novo arquivo Excel para os resultados finais
    output_file = "/Users/imac/Desktop/Esquema Inteligencia/Estudos/Resultados/Organizados/VICENTE PIRES.xlsx"
    result_workbook = openpyxl.Workbook()

    # Remover a aba padrão criada automaticamente
    default_sheet = result_workbook.active
    result_workbook.remove(default_sheet)

    # Criar uma aba para os resultados finais
    result_sheet = result_workbook.create_sheet(title="VICENTE PIRES")

    # Para cada aba no arquivo original, copiar os resultados para a nova aba no arquivo final
    for sheet_name in workbook.sheetnames:
        original_sheet = workbook[sheet_name]

        for row in original_sheet.iter_rows(values_only=True):
            result_sheet.append(row)

        result_sheet.append([])  # Adicionar duas linhas de espaço entre as tabelas

    # Salvar o novo arquivo Excel com os resultados finais no mesmo local do arquivo de origem
    result_workbook.save(output_file)
    print("Resultados finais salvos em:", output_file)

